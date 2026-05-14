# pyright: reportUndefinedVariable=false
# pylint: disable=undefined-variable
# flake8: noqa
# ruff: noqa: F821
# Transitional module executed in app.py's namespace.


def _parse_text_payload(file_bytes: bytes) -> str:
    try:
        return file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        return file_bytes.decode("latin-1")


def _extract_csv_text_from_xlsx(file_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency/runtime guard
        raise ValueError("XLSX import requires openpyxl to be installed.") from exc

    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    worksheet = workbook.active

    output = io.StringIO()
    writer = csv.writer(output)
    for row in worksheet.iter_rows(values_only=True):
        writer.writerow(["" if value is None else str(value) for value in row])

    workbook.close()
    return output.getvalue()


def split_line_preserving_coords(line: str) -> list[str]:
    """Split a CSV-ish line while preserving commas inside coordinate tuples."""
    fields: list[str] = []
    current: list[str] = []
    depth = 0
    for char in line:
        if char == "(":
            depth += 1
        elif char == ")" and depth > 0:
            depth -= 1

        if char == "," and depth == 0:
            fields.append("".join(current).strip())
            current = []
            continue

        current.append(char)

    fields.append("".join(current).strip())
    return fields


def _coerce_scrimcore_field(value: str) -> str:
    cleaned = (value or "").strip()
    if "*" in cleaned:
        return "0"
    return cleaned


def _clean_scrimcore_log_rows(rows: list[list[str]]) -> list[list[str]]:
    cleaned_rows: list[list[str]] = []
    for row in rows:
        if not row:
            continue

        if len(row) >= 2:
            event_type = (row[1] or "").strip().lower()
        else:
            event_type = (row[0] or "").strip().lower()

        if event_type == "mercy_rez":
            tail = row[2:] if len(row) >= 2 else row[1:]
            if any((field or "").strip() == "" for field in tail):
                continue

        normalized = [_coerce_scrimcore_field(field) for field in row]
        if normalized and normalized[0] == "****":
            normalized[0] = "kill"
        if len(normalized) > 1 and normalized[1] == "****":
            normalized[1] = "kill"
        cleaned_rows.append(normalized)

    return cleaned_rows


def parse_scrimcore_log_text(raw_text: str) -> tuple[dict[str, list[list[str]]], dict]:
    """Parse raw ScrimCore-style log text into categorized rows and summary metadata."""
    raw_rows = [split_line_preserving_coords(line) for line in (raw_text or "").splitlines() if (line or "").strip()]
    cleaned_rows = _clean_scrimcore_log_rows(raw_rows)

    categorized: dict[str, list[list[str]]] = defaultdict(list)
    for row in cleaned_rows:
        if len(row) < 2:
            continue

        timestamp = row[0]
        event_type = (row[1] or "").strip().lower()
        payload = row[2:]
        if not event_type:
            continue

        normalized_row = [timestamp, *payload]
        categorized[event_type].append(normalized_row)

    event_counts = {event_type: len(rows) for event_type, rows in categorized.items()}
    map_name = ""
    team1_name = ""
    team2_name = ""
    match_start_rows = categorized.get("match_start", [])
    if match_start_rows:
        first_start = match_start_rows[0]
        # match_start row layout after normalization:
        # [timestamp, match_time, map_name, mode, team1_name, team2_name]
        if len(first_start) > 2:
            map_name = str(first_start[2]).strip()
        if len(first_start) > 4:
            team1_name = str(first_start[4]).strip()
        if len(first_start) > 5:
            team2_name = str(first_start[5]).strip()

    summary = {
        "total_rows": len(cleaned_rows),
        "event_type_count": len(event_counts),
        "event_counts": dict(sorted(event_counts.items(), key=lambda item: (-item[1], item[0]))),
        "map_name": map_name,
        "team1_name": team1_name,
        "team2_name": team2_name,
    }
    return categorized, summary


def _extract_scrimcore_players_by_team(parsed_rows: dict[str, list[list[str]]]) -> dict[str, list[str]]:
    """Extract player names by parsed team label from ScrimCore event rows."""
    players_by_label: dict[str, list[str]] = defaultdict(list)
    seen_by_label: dict[str, set[str]] = defaultdict(set)

    def add_player(team_label: str, raw_player_name: str) -> None:
        label = str(team_label or "").strip()
        if not label:
            return
        if is_ringer_player_name(raw_player_name):
            return
        player_name = normalize_player_name(raw_player_name)
        if not player_name:
            return
        player_key = _compact_text(player_name) or player_name.lower()
        if player_key in seen_by_label[label]:
            return
        seen_by_label[label].add(player_key)
        players_by_label[label].append(player_name)

    # team/player field offsets by event type in parse_scrimcore_log_text output rows.
    # Each row is [timestamp, payload...], so payload starts at index 1.
    team_player_offsets: dict[str, tuple[tuple[int, int], ...]] = {
        "damage": ((2, 3), (5, 6)),
        "healing": ((2, 3), (5, 6)),
        "kill": ((2, 3), (5, 6)),
        "ability_1_used": ((2, 3),),
        "ability_2_used": ((2, 3),),
        "hero_spawn": ((2, 3),),
        "hero_swap": ((2, 3),),
        "ultimate_charged": ((2, 3),),
        "ultimate_start": ((2, 3),),
        "ultimate_end": ((2, 3),),
        "offensive_assist": ((2, 3),),
        "defensive_assist": ((2, 3),),
    }

    for event_name, rows in (parsed_rows or {}).items():
        offsets = team_player_offsets.get(str(event_name or "").strip().lower())
        if not offsets:
            continue
        for row in rows or []:
            for team_index, player_index in offsets:
                if len(row) <= player_index:
                    continue
                add_player(str(row[team_index]), str(row[player_index]))

    return {label: parse_name_list("\n".join(names)) for label, names in players_by_label.items()}


def _resolve_scrimcore_players_for_team(
    players_by_label: dict[str, list[str]],
    canonical_name: str,
    fallback_labels: list[str] | tuple[str, ...],
) -> list[str]:
    """Resolve roster names for a canonical team from parsed team labels."""
    candidate_keys = [canonical_name, *(fallback_labels or [])]
    matched_names: list[str] = []
    seen_keys: set[str] = set()

    def label_matches_candidate(label: str, candidate: str) -> bool:
        label_key = _compact_text(label)
        candidate_key = _compact_text(candidate)
        if label_key in {"team1", "team2"} or candidate_key in {"team1", "team2"}:
            return label_key == candidate_key
        return _team_names_match(label, candidate)

    for label, names in (players_by_label or {}).items():
        if not any(label_matches_candidate(label, candidate) for candidate in candidate_keys if candidate):
            continue
        for player_name in names or []:
            player_key = _compact_text(player_name) or str(player_name).strip().lower()
            if player_key and player_key not in seen_keys:
                seen_keys.add(player_key)
                matched_names.append(player_name)

    return parse_name_list("\n".join(matched_names))


def _is_scrimcore_generic_team_name(team_name: str) -> bool:
    compact = _compact_text(team_name)
    return compact in {"team1", "team2", "teamone", "teamtwo"}


def _parse_scrimcore_timestamp_seconds(raw_timestamp: str) -> int:
    text = (raw_timestamp or "").strip()
    if text.startswith("[") and text.endswith("]"):
        text = text[1:-1]
    try:
        hours, minutes, seconds = [int(part) for part in text.split(":")]
    except Exception:
        return 0
    return hours * 3600 + minutes * 60 + seconds


def _parse_scrimcore_match_time(raw_value: str | int | float | None) -> float:
    try:
        return float(str(raw_value or "0").strip() or 0)
    except Exception:
        return 0.0


def _parse_scrimcore_number(raw_value: str | int | float | None) -> float:
    try:
        return float(str(raw_value or "0").strip() or 0)
    except Exception:
        return 0.0


def _format_match_time(seconds: float) -> str:
    total = max(0, int(round(seconds)))
    return f"{total // 60:02}:{total % 60:02}"


def _resolve_scrimcore_side_slot(
    team_label: str,
    parsed_team1: str,
    parsed_team2: str,
    our_team_slot: str,
) -> str | None:
    label = str(team_label or "").strip()
    if not label:
        return None

    compact = _compact_text(label)
    if compact == "team1":
        return "team1"
    if compact == "team2":
        return "team2"

    # Draft and comp storage are keyed by map side, not by "our team".
    # ScrimCore logs label those sides as Team 1 / Team 2, so preserve that
    # mapping exactly or bans/comps land under the wrong rendered column.
    if parsed_team1 and _team_names_match(label, parsed_team1):
        return "team1"
    if parsed_team2 and _team_names_match(label, parsed_team2):
        return "team2"

    return None


def _scrimcore_round_bounds(parsed_rows: dict[str, list[list[str]]]) -> tuple[list[int], dict[int, tuple[float, float]]]:
    round_starts: dict[int, float] = {}
    for row in parsed_rows.get("round_start", []) or []:
        if len(row) <= 2:
            continue
        try:
            round_number = int(float(str(row[2]).strip()))
        except Exception:
            continue
        round_starts.setdefault(round_number, _parse_scrimcore_match_time(row[1] if len(row) > 1 else 0))

    round_ends: dict[int, float] = {}
    for row in parsed_rows.get("round_end", []) or []:
        if len(row) <= 2:
            continue
        try:
            round_number = int(float(str(row[2]).strip()))
        except Exception:
            continue
        round_ends[round_number] = _parse_scrimcore_match_time(row[1] if len(row) > 1 else 0)

    if not round_starts:
        round_starts = {1: 0.0}

    ordered_rounds = sorted(round_starts.keys())
    round_bounds: dict[int, tuple[float, float]] = {}
    for index, round_number in enumerate(ordered_rounds):
        start_time = round_starts[round_number]
        if round_number in round_ends:
            end_time = round_ends[round_number]
        elif index + 1 < len(ordered_rounds):
            end_time = max(start_time, round_starts[ordered_rounds[index + 1]] - 0.01)
        else:
            end_time = 10 ** 9
        round_bounds[round_number] = (start_time, max(start_time, end_time))

    return ordered_rounds, round_bounds


def _scrimcore_round_for_match_time(match_time: float, ordered_rounds: list[int], round_bounds: dict[int, tuple[float, float]]) -> int:
    for round_number in ordered_rounds:
        start_time, end_time = round_bounds[round_number]
        if start_time <= match_time <= end_time:
            return round_number
    candidates = [round_number for round_number in ordered_rounds if round_bounds[round_number][0] <= match_time]
    return candidates[-1] if candidates else ordered_rounds[0]


def _scrimcore_round_label(map_name: str, round_number: int) -> str:
    submaps = MAP_SUBMAPS.get(map_name, []) if map_name else []
    if 0 <= round_number - 1 < len(submaps):
        return submaps[round_number - 1]
    return f"Round {round_number}"


def _extract_scrimcore_round_scores(
    parsed_rows: dict[str, list[list[str]]],
    our_team_slot: str,
) -> tuple[str, str, dict[int, dict]]:
    round_score_rows: list[tuple[int, float, int, int, int]] = []
    for row_index, row in enumerate(parsed_rows.get("round_end", []) or []):
        if len(row) <= 5:
            continue
        try:
            round_number = int(float(str(row[2]).strip()))
            team1_score = int(float(str(row[4]).strip()))
            team2_score = int(float(str(row[5]).strip()))
        except Exception:
            continue
        round_score_rows.append((
            round_number,
            _parse_scrimcore_match_time(row[1] if len(row) > 1 else 0),
            row_index,
            team1_score,
            team2_score,
        ))

    if not round_score_rows:
        return "", "", {}

    round_score_rows.sort(key=lambda item: (item[0], item[1], item[2]))
    final_round = round_score_rows[-1]
    map_score = f"{final_round[3]}-{final_round[4]}"
    map_result = infer_result_from_score_text(map_score, slot=our_team_slot)

    round_scores: dict[int, dict] = {}
    previous_team1 = 0
    previous_team2 = 0
    for round_number, _match_time, _row_index, team1_score, team2_score in round_score_rows:
        round_team1_score = max(0, team1_score - previous_team1)
        round_team2_score = max(0, team2_score - previous_team2)
        section_score = f"{round_team1_score}-{round_team2_score}"
        round_scores[round_number] = {
            "score": section_score,
            "result": infer_result_from_score_text(section_score, slot=our_team_slot),
            "team1_score": team1_score,
            "team2_score": team2_score,
        }
        previous_team1 = team1_score
        previous_team2 = team2_score

    return map_score, map_result, round_scores


def _build_scrimcore_killfeed_events(
    parsed_rows: dict[str, list[list[str]]],
    parsed_team1: str,
    parsed_team2: str,
    our_team_slot: str,
    map_name: str,
) -> tuple[list[dict], dict]:
    ordered_rounds, round_bounds = _scrimcore_round_bounds(parsed_rows)
    kill_rows: list[dict] = []

    for index, row in enumerate(parsed_rows.get("kill", []) or []):
        if len(row) <= 7:
            continue
        match_time = _parse_scrimcore_match_time(row[1] if len(row) > 1 else 0)
        attacker_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot) or ""
        victim_slot = _resolve_scrimcore_side_slot(str(row[5]), parsed_team1, parsed_team2, our_team_slot) or ""
        round_number = _scrimcore_round_for_match_time(match_time, ordered_rounds, round_bounds)
        kill_rows.append(
            {
                "index": index,
                "timestamp": _format_match_time(match_time),
                "match_time": match_time,
                "round_number": round_number,
                "round_label": _scrimcore_round_label(map_name, round_number),
                "attacker_team": attacker_slot,
                "attacker_name": normalize_player_name(str(row[3])),
                "attacker_hero": normalize_hero_slot_value(str(row[4])),
                "victim_team": victim_slot,
                "victim_name": normalize_player_name(str(row[6])),
                "victim_hero": normalize_hero_slot_value(str(row[7])),
                "ability": normalize_hero_slot_value(str(row[8])) if len(row) > 8 else "",
            }
        )

    kill_rows.sort(key=lambda item: (item["match_time"], item["index"]))

    fights: list[list[dict]] = []
    current_fight: list[dict] = []
    current_end = 0.0
    for kill in kill_rows:
        if not current_fight or kill["match_time"] - current_end > 15:
            current_fight = [kill]
            fights.append(current_fight)
        else:
            current_fight.append(kill)
        current_end = kill["match_time"]

    events: list[dict] = []
    first_deaths_by_team = {"team1": 0, "team2": 0}
    first_deaths_by_player: Counter[str] = Counter()
    fight_wins_by_team = {"team1": 0, "team2": 0}
    fight_rows: list[dict] = []

    for fight_number, fight in enumerate(fights, start=1):
        first_kill = fight[0] if fight else None
        fight_kills_by_team = Counter(kill.get("attacker_team") for kill in fight)
        fight_winner = ""
        if fight_kills_by_team["team1"] != fight_kills_by_team["team2"]:
            fight_winner = "team1" if fight_kills_by_team["team1"] > fight_kills_by_team["team2"] else "team2"
            fight_wins_by_team[fight_winner] += 1
        if fight:
            fight_rows.append({
                "fight_number": fight_number,
                "start": fight[0]["match_time"],
                "end": fight[-1]["match_time"],
                "start_label": _format_match_time(fight[0]["match_time"]),
                "end_label": _format_match_time(fight[-1]["match_time"]),
                "winner": fight_winner,
                "team1_kills": fight_kills_by_team["team1"],
                "team2_kills": fight_kills_by_team["team2"],
                "round_label": fight[0].get("round_label", ""),
            })
        for kill in fight:
            is_first = kill is first_kill
            if is_first and kill["victim_team"] in first_deaths_by_team:
                first_deaths_by_team[kill["victim_team"]] += 1
                if kill["victim_name"]:
                    first_deaths_by_player[kill["victim_name"]] += 1

            event_type = "First Kill" if is_first else "Pick"
            ability = kill.get("ability") or "Unknown"
            description = (
                f"{kill['attacker_name']} ({kill['attacker_hero'] or 'Unknown'}) "
                f"eliminated {kill['victim_name']} ({kill['victim_hero'] or 'Unknown'})"
                f" with {ability}."
            )
            events.append(
                {
                    "id": 0,
                    "timestamp": kill["timestamp"],
                    "match_time": kill["match_time"],
                    "event_type": event_type,
                    "description": description,
                    "event_source": "scrimcore_killfeed",
                    "fight_number": fight_number,
                    "round_number": kill["round_number"],
                    "round_label": kill["round_label"],
                    "attacker_team": kill["attacker_team"],
                    "attacker_name": kill["attacker_name"],
                    "attacker_hero": kill["attacker_hero"],
                    "victim_team": kill["victim_team"],
                    "victim_name": kill["victim_name"],
                    "victim_hero": kill["victim_hero"],
                    "ability": ability,
                    "is_first_kill": is_first,
                }
            )

            if is_first:
                events.append(
                    {
                        "id": 0,
                        "timestamp": kill["timestamp"],
                        "match_time": kill["match_time"],
                        "event_type": "First Death",
                        "description": (
                            f"Fight {fight_number}: {kill['victim_name']} "
                            f"({kill['victim_hero'] or 'Unknown'}) died first."
                        ),
                        "event_source": "scrimcore_first_death_marker",
                        "fight_number": fight_number,
                        "round_number": kill["round_number"],
                        "round_label": kill["round_label"],
                        "victim_team": kill["victim_team"],
                        "victim_name": kill["victim_name"],
                        "victim_hero": kill["victim_hero"],
                    }
                )

    events.sort(key=lambda item: (
        float(item.get("match_time") or 0),
        1 if item.get("event_source") == "scrimcore_first_death_marker" else 0,
    ))

    summary = {
        "kill_count": len(kill_rows),
        "fight_count": len(fights),
        "team1_fight_wins": fight_wins_by_team["team1"],
        "team2_fight_wins": fight_wins_by_team["team2"],
        "fights": fight_rows,
        "team1_first_deaths": first_deaths_by_team["team1"],
        "team2_first_deaths": first_deaths_by_team["team2"],
        "top_first_deaths": [
            {"player": player, "count": count}
            for player, count in first_deaths_by_player.most_common(5)
        ],
    }
    return events, summary


def _build_scrimcore_analysis_events(
    parsed_rows: dict[str, list[list[str]]],
    parsed_team1: str,
    parsed_team2: str,
    our_team_slot: str,
    map_name: str,
) -> tuple[list[dict], dict]:
    ordered_rounds, round_bounds = _scrimcore_round_bounds(parsed_rows)

    def row_match_time(row: list[str]) -> float:
        return _parse_scrimcore_match_time(row[1] if len(row) > 1 else 0)

    def row_round_label(match_time: float) -> tuple[int, str]:
        round_number = _scrimcore_round_for_match_time(match_time, ordered_rounds, round_bounds)
        return round_number, _scrimcore_round_label(map_name, round_number)

    def side_name(side_slot: str) -> str:
        if side_slot == "team1":
            return parsed_team1 or "Team 1"
        if side_slot == "team2":
            return parsed_team2 or "Team 2"
        return "Unknown Team"

    events: list[dict] = []
    team_counts_template = {"team1": 0, "team2": 0}
    ult_counts = dict(team_counts_template)
    swap_counts = dict(team_counts_template)
    ability_counts = dict(team_counts_template)
    objective_counts = dict(team_counts_template)
    ult_players: Counter[str] = Counter()
    swap_players: Counter[str] = Counter()
    swap_pairs: Counter[str] = Counter()
    ability_players: Counter[str] = Counter()
    first_ult_by_team: dict[str, dict] = {}
    first_ability_by_team: dict[str, dict] = {}
    ultimate_spans_by_player: dict[tuple[str, str], list[dict]] = defaultdict(list)

    for event_name in ("ultimate_start", "ultimate_end", "ultimate_charged"):
        for row_index, row in enumerate(parsed_rows.get(event_name, []) or []):
            if len(row) <= 4:
                continue
            match_time = row_match_time(row)
            side_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot) or ""
            player_name = normalize_player_name(str(row[3]))
            hero_name = normalize_hero_slot_value(str(row[4]))
            round_number, round_label = row_round_label(match_time)
            label = "Ult Used" if event_name == "ultimate_start" else ("Ult Ended" if event_name == "ultimate_end" else "Ult Charged")
            if event_name == "ultimate_start":
                if side_slot in ult_counts:
                    ult_counts[side_slot] += 1
                if player_name:
                    ult_players[f"{player_name} ({hero_name or 'Unknown'})"] += 1
                    ultimate_spans_by_player[(side_slot, player_name)].append({
                        "start": match_time,
                        "end": match_time + 10,
                        "player": player_name,
                        "hero": hero_name,
                        "team": side_slot,
                    })
                first_ult_by_team.setdefault(side_slot, {
                    "player": player_name,
                    "hero": hero_name,
                    "time": _format_match_time(match_time),
                    "round_label": round_label,
                })
            events.append({
                "id": 0,
                "timestamp": _format_match_time(match_time),
                "match_time": match_time,
                "event_type": label,
                "description": f"{player_name} ({hero_name or 'Unknown'}) {label.lower()} for {side_name(side_slot)}.",
                "event_source": f"scrimcore_{event_name}",
                "round_number": round_number,
                "round_label": round_label,
                "team": side_slot,
                "player_name": player_name,
                "hero": hero_name,
                "sort_hint": row_index,
            })

    for row in parsed_rows.get("ultimate_end", []) or []:
        if len(row) <= 4:
            continue
        match_time = row_match_time(row)
        side_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot) or ""
        player_name = normalize_player_name(str(row[3]))
        spans = ultimate_spans_by_player.get((side_slot, player_name), [])
        open_spans = [span for span in spans if span["start"] <= match_time and span["end"] == span["start"] + 10]
        if open_spans:
            open_spans[-1]["end"] = match_time

    for row_index, row in enumerate(parsed_rows.get("hero_swap", []) or []):
        if len(row) <= 5:
            continue
        match_time = row_match_time(row)
        side_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot) or ""
        player_name = normalize_player_name(str(row[3]))
        hero_name = normalize_hero_slot_value(str(row[4]))
        previous_hero = normalize_hero_slot_value(str(row[5]))
        round_number, round_label = row_round_label(match_time)
        if side_slot in swap_counts:
            swap_counts[side_slot] += 1
        if player_name:
            swap_players[player_name] += 1
        if previous_hero or hero_name:
            swap_pairs[f"{previous_hero or 'Unknown'} -> {hero_name or 'Unknown'}"] += 1
        events.append({
            "id": 0,
            "timestamp": _format_match_time(match_time),
            "match_time": match_time,
            "event_type": "Hero Swap",
            "description": f"{player_name} swapped from {previous_hero or 'Unknown'} to {hero_name or 'Unknown'}.",
            "event_source": "scrimcore_hero_swap",
            "round_number": round_number,
            "round_label": round_label,
            "team": side_slot,
            "player_name": player_name,
            "hero": hero_name,
            "previous_hero": previous_hero,
            "sort_hint": row_index,
        })

    for event_name in ("ability_1_used", "ability_2_used"):
        ability_label = "Ability 1" if event_name == "ability_1_used" else "Ability 2"
        for row in parsed_rows.get(event_name, []) or []:
            if len(row) <= 4:
                continue
            match_time = row_match_time(row)
            side_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot) or ""
            player_name = normalize_player_name(str(row[3]))
            hero_name = normalize_hero_slot_value(str(row[4]))
            round_number, round_label = row_round_label(match_time)
            if side_slot in ability_counts:
                ability_counts[side_slot] += 1
            if player_name:
                ability_players[f"{player_name} ({hero_name or 'Unknown'})"] += 1
            first_ability_by_team.setdefault(side_slot, {
                "player": player_name,
                "hero": hero_name,
                "ability": ability_label,
                "time": _format_match_time(match_time),
                "round_label": round_label,
            })

    for row_index, row in enumerate(parsed_rows.get("objective_captured", []) or []):
        if len(row) <= 3:
            continue
        match_time = row_match_time(row)
        side_slot = _resolve_scrimcore_side_slot(str(row[3]), parsed_team1, parsed_team2, our_team_slot) or ""
        round_number, round_label = row_round_label(match_time)
        if side_slot in objective_counts:
            objective_counts[side_slot] += 1
        events.append({
            "id": 0,
            "timestamp": _format_match_time(match_time),
            "match_time": match_time,
            "event_type": "Objective Taken",
            "description": f"{side_name(side_slot)} captured objective {str(row[4]) if len(row) > 4 else ''}.".strip(),
            "event_source": "scrimcore_objective_captured",
            "round_number": round_number,
            "round_label": round_label,
            "team": side_slot,
            "sort_hint": row_index,
        })

    for event_name, label in (("round_start", "Round Start"), ("round_end", "Round End"), ("match_start", "Match Start"), ("match_end", "Match End")):
        for row_index, row in enumerate(parsed_rows.get(event_name, []) or []):
            match_time = row_match_time(row)
            round_number, round_label = row_round_label(match_time)
            events.append({
                "id": 0,
                "timestamp": _format_match_time(match_time),
                "match_time": match_time,
                "event_type": label,
                "description": f"{label}{f' - {round_label}' if event_name in {'round_start', 'round_end'} else ''}.",
                "event_source": f"scrimcore_{event_name}",
                "round_number": round_number,
                "round_label": round_label,
                "sort_hint": row_index,
            })

    events.sort(key=lambda item: (float(item.get("match_time") or 0), str(item.get("event_source") or ""), int(item.get("sort_hint") or 0)))

    killfeed_events, killfeed_summary = _build_scrimcore_killfeed_events(parsed_rows, parsed_team1, parsed_team2, our_team_slot, map_name)
    team_kills = {"team1": 0, "team2": 0}
    team_deaths = {"team1": 0, "team2": 0}
    ultimate_value = {"team1": 0, "team2": 0}
    ultimate_value_players: Counter[str] = Counter()
    rotation_deaths_by_team = {"team1": 0, "team2": 0}
    rotation_death_players: Counter[str] = Counter()
    rotation_death_events: list[dict] = []
    ability_phase_labels = {
        "pre_fight": "Pre-Fight",
        "early": "Early",
        "mid": "Mid",
        "late": "Late",
        "cleanup": "Cleanup",
    }
    ability_phase_stats = {
        "team1": defaultdict(lambda: {
            "hero": "",
            "total": 0,
            "phases": {phase: {"count": 0, "wins": 0} for phase in ability_phase_labels.keys()},
        }),
        "team2": defaultdict(lambda: {
            "hero": "",
            "total": 0,
            "phases": {phase: {"count": 0, "wins": 0} for phase in ability_phase_labels.keys()},
        }),
    }
    ability_phase_seen: dict[str, set[tuple[int, str, str]]] = {"team1": set(), "team2": set()}

    allowed_source_types = {"ability", "ability_1", "ability_2", "ability_3", "ultimate_ability"}
    allowed_ability_names_by_hero: dict[str, set[str]] = defaultdict(set)
    canonical_ability_name_by_hero_source: dict[str, dict[str, str]] = defaultdict(dict)
    global_allowed_ability_names: set[str] = {"ultimate", "ult"}
    try:
        from hero_ability_details import HERO_ABILITY_DETAILS

        for raw_hero_name, rows in (HERO_ABILITY_DETAILS or {}).items():
            hero_key = _compact_text(normalize_hero_slot_value(str(raw_hero_name)))
            if not hero_key:
                continue
            for row in rows or []:
                source_type = str(row.get("source_type") or "").strip().lower()
                if source_type not in allowed_source_types:
                    continue
                canonical_ability = normalize_hero_slot_value(str(row.get("name") or ""))
                ability_key = _compact_text(canonical_ability)
                if not ability_key:
                    continue
                allowed_ability_names_by_hero[hero_key].add(ability_key)
                global_allowed_ability_names.add(ability_key)
                canonical_ability_name_by_hero_source[hero_key].setdefault(source_type, canonical_ability)
    except Exception:
        pass
    kill_times: list[float] = []
    fight_map = {
        int(fight.get("fight_number") or 0): fight
        for fight in (killfeed_summary.get("fights") or [])
        if int(fight.get("fight_number") or 0) > 0
    }

    def _ability_phase_for_offset(seconds_since_fight_start: float) -> str:
        if seconds_since_fight_start <= 3:
            return "pre_fight"
        if seconds_since_fight_start <= 8:
            return "early"
        if seconds_since_fight_start <= 14:
            return "mid"
        if seconds_since_fight_start <= 20:
            return "late"
        return "cleanup"

    for event in killfeed_events:
        if event.get("event_source") != "scrimcore_killfeed":
            continue
        attacker_team = event.get("attacker_team")
        victim_team = event.get("victim_team")
        if attacker_team in team_kills:
            team_kills[attacker_team] += 1
        if victim_team in team_deaths:
            team_deaths[victim_team] += 1
        kill_times.append(float(event.get("match_time") or 0))
        fight_number = int(event.get("fight_number") or 0)
        fight_meta = fight_map.get(fight_number, {})
        # Treat one-kill fights as rotation picks; this keeps the signal explicit
        # without requiring additional parser-only events.
        if fight_meta.get("team1_kills", 0) + fight_meta.get("team2_kills", 0) == 1 and victim_team in rotation_deaths_by_team:
            rotation_deaths_by_team[victim_team] += 1
            victim_name = str(event.get("victim_name") or "")
            if victim_name:
                rotation_death_players[victim_name] += 1
            rotation_death_events.append(
                {
                    "timestamp": event.get("timestamp") or "",
                    "round_label": event.get("round_label") or "",
                    "victim_team": victim_team,
                    "victim_name": victim_name,
                    "victim_hero": event.get("victim_hero") or "",
                    "attacker_name": event.get("attacker_name") or "",
                    "attacker_hero": event.get("attacker_hero") or "",
                }
            )

        kill_time = float(event.get("match_time") or 0)
        attacker_name = str(event.get("attacker_name") or "")
        for span in ultimate_spans_by_player.get((str(attacker_team or ""), attacker_name), []):
            if span["start"] <= kill_time <= span["end"]:
                if attacker_team in ultimate_value:
                    ultimate_value[attacker_team] += 1
                ultimate_value_players[f"{attacker_name} ({event.get('attacker_hero') or 'Unknown'})"] += 1
                break

    fight_windows: list[dict] = []
    for fight in (killfeed_summary.get("fights") or []):
        fight_number = int(fight.get("fight_number") or 0)
        if fight_number <= 0:
            continue
        fight_windows.append(
            {
                "fight_number": fight_number,
                "start": float(fight.get("start") or 0),
                "end": float(fight.get("end") or 0),
                "winner": str(fight.get("winner") or ""),
            }
        )

    def _match_fight_for_ability_time(match_time: float) -> dict | None:
        # Include a pre-fight and post-fight leeway so setup abilities in losing fights are represented.
        pre_fight_leeway = 8.0
        post_fight_leeway = 5.0
        best_fight = None
        best_distance = None
        for fight in fight_windows:
            start = fight["start"]
            end = fight["end"]
            if match_time < (start - pre_fight_leeway) or match_time > (end + post_fight_leeway):
                continue
            distance = abs(match_time - start)
            if best_distance is None or distance < best_distance:
                best_distance = distance
                best_fight = fight
        return best_fight

    ability_event_specs = (
        ("ability_1_used", "ability_1"),
        ("ability_2_used", "ability_2"),
        ("ultimate_start", "ultimate_ability"),
    )

    for event_name, source_type in ability_event_specs:
        for row in parsed_rows.get(event_name, []) or []:
            if len(row) <= 4:
                continue

            match_time = _parse_scrimcore_match_time(row[1] if len(row) > 1 else 0)
            side_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot) or ""
            if side_slot not in ability_phase_stats:
                continue

            player_name = normalize_player_name(str(row[3]))
            hero_name = normalize_hero_slot_value(str(row[4]))
            hero_key = _compact_text(hero_name)

            canonical_ability = canonical_ability_name_by_hero_source.get(hero_key, {}).get(source_type, "")
            if not canonical_ability:
                canonical_ability = "Ultimate" if source_type == "ultimate_ability" else ("Ability 1" if source_type == "ability_1" else "Ability 2")

            ability_name = normalize_hero_slot_value(canonical_ability)
            ability_key = _compact_text(ability_name)
            hero_allowed_names = allowed_ability_names_by_hero.get(hero_key, set())
            is_whitelisted_ability = bool(
                ability_key
                and (
                    ability_key in hero_allowed_names
                    or ability_key in global_allowed_ability_names
                )
            )
            if not is_whitelisted_ability:
                continue

            matched_fight = _match_fight_for_ability_time(match_time)
            if matched_fight:
                fight_number = int(matched_fight.get("fight_number") or 0)
                fight_start = float(matched_fight.get("start") or match_time)
                fight_winner = str(matched_fight.get("winner") or "")
            else:
                fight_number = 0
                fight_start = match_time
                fight_winner = ""

            phase_key = _ability_phase_for_offset(max(0.0, match_time - fight_start))
            dedupe_key = (
                fight_number,
                player_name or "unknown-player",
                ability_name,
                phase_key,
            )
            if dedupe_key in ability_phase_seen[side_slot]:
                continue

            ability_phase_seen[side_slot].add(dedupe_key)
            ability_row = ability_phase_stats[side_slot][ability_name]
            ability_row["total"] += 1
            if not ability_row["hero"]:
                ability_row["hero"] = hero_name
            ability_row["phases"][phase_key]["count"] += 1
            if fight_winner == side_slot:
                ability_row["phases"][phase_key]["wins"] += 1

    player_stat_rows: list[dict] = []
    team_stat_totals = {
        "team1": {"hero_damage": 0.0, "healing": 0.0, "damage_taken": 0.0, "time_played": 0.0, "eliminations": 0, "deaths": 0, "ults_used": 0},
        "team2": {"hero_damage": 0.0, "healing": 0.0, "damage_taken": 0.0, "time_played": 0.0, "eliminations": 0, "deaths": 0, "ults_used": 0},
    }
    damage_by_round = defaultdict(lambda: {"team1": 0.0, "team2": 0.0, "round_label": ""})
    for row in parsed_rows.get("player_stat", []) or []:
        if len(row) <= 38:
            continue
        side_slot = _resolve_scrimcore_side_slot(str(row[3]), parsed_team1, parsed_team2, our_team_slot) or ""
        stat_match_time = _parse_scrimcore_match_time(row[1] if len(row) > 1 else 0)
        stat_round_number, stat_round_label = row_round_label(stat_match_time)
        player_name = normalize_player_name(str(row[4]))
        hero_name = normalize_hero_slot_value(str(row[5]))
        stat_row = {
            "team": side_slot,
            "player": player_name,
            "hero": hero_name,
            "eliminations": int(_parse_scrimcore_number(row[6])),
            "final_blows": int(_parse_scrimcore_number(row[7])),
            "deaths": int(_parse_scrimcore_number(row[8])),
            "hero_damage": round(_parse_scrimcore_number(row[11]), 2),
            "healing": round(_parse_scrimcore_number(row[12]), 2),
            "damage_taken": round(_parse_scrimcore_number(row[15]), 2),
            "ultimates_earned": int(_parse_scrimcore_number(row[19])),
            "ultimates_used": int(_parse_scrimcore_number(row[20])),
            "multikills": int(_parse_scrimcore_number(row[22])),
            "hero_time_played": round(_parse_scrimcore_number(row[38]), 2),
        }
        player_stat_rows.append(stat_row)
        if side_slot in team_stat_totals:
            team_stat_totals[side_slot]["hero_damage"] += stat_row["hero_damage"]
            team_stat_totals[side_slot]["healing"] += stat_row["healing"]
            team_stat_totals[side_slot]["damage_taken"] += stat_row["damage_taken"]
            team_stat_totals[side_slot]["time_played"] += stat_row["hero_time_played"]
            team_stat_totals[side_slot]["eliminations"] += stat_row["eliminations"]
            team_stat_totals[side_slot]["deaths"] += stat_row["deaths"]
            team_stat_totals[side_slot]["ults_used"] += stat_row["ultimates_used"]
            damage_by_round[stat_round_number][side_slot] += stat_row["hero_damage"]
            damage_by_round[stat_round_number]["round_label"] = stat_round_label

    for side_slot in ("team1", "team2"):
        for key, value in list(team_stat_totals[side_slot].items()):
            if isinstance(value, float):
                team_stat_totals[side_slot][key] = round(value, 2)

    ability_timing_tables: dict[str, list[dict]] = {"team1": [], "team2": []}
    ability_timing_insights: dict[str, list[dict]] = {"team1": [], "team2": []}
    filtered_team_totals = {"team1": 0, "team2": 0}
    for team_slot in ("team1", "team2"):
        sorted_abilities = sorted(
            ability_phase_stats[team_slot].items(),
            key=lambda item: (int(item[1].get("total") or 0), item[0]),
            reverse=True,
        )
        filtered_team_totals[team_slot] = sum(int(item[1].get("total") or 0) for item in sorted_abilities)
        top_rows = sorted_abilities[:10]
        for ability_name, stat_row in top_rows:
            phase_counts = {}
            phase_winrates = {}
            for phase_key in ability_phase_labels.keys():
                count = int(stat_row["phases"][phase_key]["count"])
                wins = int(stat_row["phases"][phase_key]["wins"])
                phase_counts[phase_key] = count
                # Display any observed sample while smoothing small-sample volatility.
                phase_winrates[phase_key] = round(((wins + 1) / (count + 2)) * 100) if count else None

            ability_timing_tables[team_slot].append(
                {
                    "ability": ability_name,
                    "hero": stat_row.get("hero", ""),
                    "total": int(stat_row.get("total") or 0),
                    "phase_counts": phase_counts,
                    "phase_winrate": phase_winrates,
                }
            )

        insight_candidates: list[tuple[int, int, str, str]] = []
        for ability_name, stat_row in sorted_abilities:
            for phase_key in ability_phase_labels.keys():
                count = int(stat_row["phases"][phase_key]["count"])
                if count < 3:
                    continue
                wins = int(stat_row["phases"][phase_key]["wins"])
                winrate = round(((wins + 1) / (count + 2)) * 100)
                insight_candidates.append((winrate, count, ability_name, phase_key))

        insight_candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)

        # Keep insight cards diverse: prefer one card per unique ability first.
        selected_insights: list[tuple[int, int, str, str]] = []
        seen_abilities: set[str] = set()
        for candidate in insight_candidates:
            ability_name = candidate[2]
            if ability_name in seen_abilities:
                continue
            selected_insights.append(candidate)
            seen_abilities.add(ability_name)
            if len(selected_insights) >= 3:
                break

        # If there are fewer than 3 unique abilities, fill with best remaining candidates.
        if len(selected_insights) < 3:
            for candidate in insight_candidates:
                if candidate in selected_insights:
                    continue
                selected_insights.append(candidate)
                if len(selected_insights) >= 3:
                    break

        for winrate, count, ability_name, phase_key in selected_insights:
            if phase_key == "pre_fight":
                pattern = "strong initiation pattern"
            elif phase_key == "early":
                pattern = "strong timing pattern"
            elif phase_key == "mid":
                pattern = "mid-fight conversion pattern"
            elif phase_key == "late":
                pattern = "late-fight closure pattern"
            else:
                pattern = "cleanup conversion pattern"
            ability_timing_insights[team_slot].append(
                {
                    "ability": ability_name,
                    "phase": ability_phase_labels[phase_key],
                    "winrate": winrate,
                    "samples": count,
                    "pattern": pattern,
                }
            )

    damage_round_rows: list[dict] = []
    cumulative_team1 = 0.0
    cumulative_team2 = 0.0
    for round_number in sorted(damage_by_round.keys()):
        round_team1 = round(float(damage_by_round[round_number]["team1"]), 2)
        round_team2 = round(float(damage_by_round[round_number]["team2"]), 2)
        cumulative_team1 = round(cumulative_team1 + round_team1, 2)
        cumulative_team2 = round(cumulative_team2 + round_team2, 2)
        damage_round_rows.append({
            "round_number": round_number,
            "round_label": damage_by_round[round_number]["round_label"] or _scrimcore_round_label(map_name, round_number),
            "team1_damage": round_team1,
            "team2_damage": round_team2,
            "team1_cumulative": cumulative_team1,
            "team2_cumulative": cumulative_team2,
        })

    summary = {
        "first_deaths": killfeed_summary,
        "ultimates": {
            "team1_count": ult_counts["team1"],
            "team2_count": ult_counts["team2"],
            "total": ult_counts["team1"] + ult_counts["team2"],
            "team1_value_kills": ultimate_value["team1"],
            "team2_value_kills": ultimate_value["team2"],
            "top_value_players": [{"label": label, "count": count} for label, count in ultimate_value_players.most_common(5)],
            "first_by_team": first_ult_by_team,
            "top_players": [{"label": label, "count": count} for label, count in ult_players.most_common(5)],
        },
        "timing": {
            "first_kill_time": _format_match_time(min(kill_times)) if kill_times else "",
            "last_kill_time": _format_match_time(max(kill_times)) if kill_times else "",
            "fight_count": killfeed_summary.get("fight_count", 0),
            "kill_count": killfeed_summary.get("kill_count", 0),
        },
        "efficiency": {
            "team1_kills": team_kills["team1"],
            "team2_kills": team_kills["team2"],
            "team1_deaths": team_deaths["team1"],
            "team2_deaths": team_deaths["team2"],
            "team1_diff": team_kills["team1"] - team_deaths["team1"],
            "team2_diff": team_kills["team2"] - team_deaths["team2"],
            "team1_objectives": objective_counts["team1"],
            "team2_objectives": objective_counts["team2"],
        },
        "swaps": {
            "team1_count": swap_counts["team1"],
            "team2_count": swap_counts["team2"],
            "total": swap_counts["team1"] + swap_counts["team2"],
            "top_swappers": [{"label": label, "count": count} for label, count in swap_players.most_common(5)],
            "top_pairs": [{"label": label, "count": count} for label, count in swap_pairs.most_common(5)],
        },
        "rotation_deaths": {
            "team1_count": rotation_deaths_by_team["team1"],
            "team2_count": rotation_deaths_by_team["team2"],
            "total": rotation_deaths_by_team["team1"] + rotation_deaths_by_team["team2"],
            "top_victims": [{"player": player, "count": count} for player, count in rotation_death_players.most_common(5)],
            "events": rotation_death_events[:20],
        },
        "objectives": {
            "team1_count": objective_counts["team1"],
            "team2_count": objective_counts["team2"],
            "total": objective_counts["team1"] + objective_counts["team2"],
        },
        "ability_timing": {
            "team1_count": filtered_team_totals["team1"],
            "team2_count": filtered_team_totals["team2"],
            "total": filtered_team_totals["team1"] + filtered_team_totals["team2"],
            "first_by_team": first_ability_by_team,
            "top_players": [{"label": label, "count": count} for label, count in ability_players.most_common(5)],
            "phase_labels": ability_phase_labels,
            "team_tables": ability_timing_tables,
            "insights": ability_timing_insights,
        },
        "player_stats": {
            "team_totals": team_stat_totals,
            "top_damage": sorted(player_stat_rows, key=lambda item: item["hero_damage"], reverse=True)[:8],
            "top_healing": sorted(player_stat_rows, key=lambda item: item["healing"], reverse=True)[:8],
            "rows": player_stat_rows,
        },
        "damage_by_round": {
            "rows": damage_round_rows,
        },
    }
    return events, summary


def _extract_scrimcore_draft(
    parsed_rows: dict[str, list[list[str]]],
    parsed_team1: str,
    parsed_team2: str,
    our_team_slot: str,
) -> dict:
    draft = {
        "team1": {"ban1": "", "ban2": "", "protect1": "", "ban3": "", "protect2": "", "ban4": ""},
        "team2": {"ban1": "", "ban2": "", "protect1": "", "ban3": "", "protect2": "", "ban4": ""},
    }
    ban_counts = {"team1": 0, "team2": 0}
    protect_counts = {"team1": 0, "team2": 0}

    for row in parsed_rows.get("hero_ban", []) or []:
        if len(row) <= 4:
            continue
        slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot)
        hero_name = normalize_hero_slot_value(str(row[4]))
        if not slot or not hero_name:
            continue
        if ban_counts[slot] >= 4:
            continue
        ban_counts[slot] += 1
        draft[slot][f"ban{ban_counts[slot]}"] = hero_name

    for row in parsed_rows.get("hero_protect", []) or []:
        if len(row) <= 3:
            continue
        slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot)
        hero_name = normalize_hero_slot_value(str(row[3]))
        if not slot or not hero_name:
            continue
        if protect_counts[slot] >= 2:
            continue
        protect_counts[slot] += 1
        draft[slot][f"protect{protect_counts[slot]}"] = hero_name

    return draft


def _extract_scrimcore_comp_sections(
    parsed_rows: dict[str, list[list[str]]],
    parsed_team1: str,
    parsed_team2: str,
    our_team_slot: str,
    map_name: str = "",
) -> list[dict]:
    _map_score, _map_result, round_scores = _extract_scrimcore_round_scores(parsed_rows, our_team_slot)
    round_starts: dict[int, int] = {}
    for row in parsed_rows.get("round_start", []) or []:
        if len(row) <= 2:
            continue
        try:
            round_number = int(float(str(row[2]).strip()))
        except Exception:
            continue
        round_starts.setdefault(round_number, _parse_scrimcore_timestamp_seconds(str(row[0])))

    round_ends: dict[int, int] = {}
    for row in parsed_rows.get("round_end", []) or []:
        if len(row) <= 2:
            continue
        try:
            round_number = int(float(str(row[2]).strip()))
        except Exception:
            continue
        round_ends[round_number] = _parse_scrimcore_timestamp_seconds(str(row[0]))

    if not round_starts:
        round_starts = {1: 0}

    ordered_rounds = sorted(round_starts.keys())
    round_bounds: dict[int, tuple[int, int]] = {}
    for index, round_number in enumerate(ordered_rounds):
        start_ts = round_starts[round_number]
        if round_number in round_ends:
            end_ts = round_ends[round_number]
        elif index + 1 < len(ordered_rounds):
            end_ts = round_starts[ordered_rounds[index + 1]] - 1
        else:
            end_ts = 10 ** 9
        round_bounds[round_number] = (start_ts, max(start_ts, end_ts))

    player_heroes: dict[int, dict[str, dict[str, str]]] = {
        round_number: {"team1": {}, "team2": {}}
        for round_number in ordered_rounds
    }
    player_order: dict[int, dict[str, list[str]]] = {
        round_number: {"team1": [], "team2": []}
        for round_number in ordered_rounds
    }

    def round_for_timestamp(ts_seconds: int) -> int:
        for round_number in ordered_rounds:
            start_ts, end_ts = round_bounds[round_number]
            if start_ts <= ts_seconds <= end_ts:
                return round_number
        candidates = [round_number for round_number in ordered_rounds if round_bounds[round_number][0] <= ts_seconds]
        return candidates[-1] if candidates else ordered_rounds[0]

    def set_player(round_number: int, side_slot: str, raw_player: str, raw_hero: str) -> None:
        if side_slot not in ("team1", "team2"):
            return
        if is_ringer_player_name(raw_player):
            return
        player_name = normalize_player_name(raw_player)
        hero_name = normalize_hero_slot_value(raw_hero)
        if not player_name or not hero_name:
            return

        if player_name not in player_heroes[round_number][side_slot]:
            if len(player_order[round_number][side_slot]) >= 6:
                return
            player_order[round_number][side_slot].append(player_name)
        player_heroes[round_number][side_slot][player_name] = hero_name

    event_rows: list[tuple[int, int, str, list[str]]] = []
    event_index = 0
    tracked_events = {
        "hero_spawn", "hero_swap", "damage", "healing", "kill",
    }
    for event_name in tracked_events:
        for row in parsed_rows.get(event_name, []) or []:
            event_rows.append((_parse_scrimcore_timestamp_seconds(str(row[0])), event_index, event_name, row))
            event_index += 1
    event_rows.sort(key=lambda item: (item[0], item[1]))

    for ts_seconds, _idx, event_name, row in event_rows:
        round_number = round_for_timestamp(ts_seconds)
        if event_name in {"hero_spawn", "hero_swap"}:
            if len(row) <= 4:
                continue
            side_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot)
            set_player(round_number, str(side_slot or ""), str(row[3]), str(row[4]))
            continue

        if event_name in {"damage", "healing", "kill"}:
            if len(row) > 4:
                atk_slot = _resolve_scrimcore_side_slot(str(row[2]), parsed_team1, parsed_team2, our_team_slot)
                set_player(round_number, str(atk_slot or ""), str(row[3]), str(row[4]))
            if len(row) > 7:
                vic_slot = _resolve_scrimcore_side_slot(str(row[5]), parsed_team1, parsed_team2, our_team_slot)
                set_player(round_number, str(vic_slot or ""), str(row[6]), str(row[7]))

    submaps = MAP_SUBMAPS.get(map_name, []) if map_name else []
    is_attack_defense_map = map_name in ATTACK_DEFENSE_MAPS
    comp_sections: list[dict] = []
    for round_number in ordered_rounds:
        section = {
            "submap": submaps[round_number - 1] if 0 <= round_number - 1 < len(submaps) else f"Round {round_number}",
            "side": "Attack" if is_attack_defense_map and round_number % 2 == 1 else ("Defense" if is_attack_defense_map else ""),
            "score": round_scores.get(round_number, {}).get("score", ""),
            "result": round_scores.get(round_number, {}).get("result", ""),
            "team1": [],
            "team2": [],
        }
        for side_slot in ("team1", "team2"):
            for player_name in player_order[round_number][side_slot]:
                section[side_slot].append(
                    {
                        "player": player_name,
                        "hero": player_heroes[round_number][side_slot].get(player_name, ""),
                    }
                )
            while len(section[side_slot]) < 6:
                section[side_slot].append({"player": "", "hero": ""})
        comp_sections.append(section)

    return comp_sections


def _build_round_hero_summary(comp_sections: list[dict]) -> list[dict]:
    summaries: list[dict] = []
    for index, section in enumerate(comp_sections or [], start=1):
        counts: Counter[str] = Counter()
        team_counts = {"team1": Counter(), "team2": Counter()}
        for side in ("team1", "team2"):
            for slot in section.get(side, []) or []:
                hero_name = normalize_hero_slot_value(slot.get("hero", ""))
                if not hero_name:
                    continue
                counts[hero_name] += 1
                team_counts[side][hero_name] += 1

        summaries.append(
            {
                "round_number": index,
                "round_label": section.get("submap") or f"Round {index}",
                "side": section.get("side", ""),
                "top_heroes": [
                    {"hero": hero, "count": count}
                    for hero, count in counts.most_common(8)
                ],
                "team1_heroes": [
                    {"hero": hero, "count": count}
                    for hero, count in team_counts["team1"].most_common(6)
                ],
                "team2_heroes": [
                    {"hero": hero, "count": count}
                    for hero, count in team_counts["team2"].most_common(6)
                ],
            }
        )
    return summaries


def _cache_only_log_import_mode() -> bool:
    """Return True when experiment log imports should skip DB persistence."""
    raw = (
        os.environ.get("PREVIEW_CACHE_ONLY_LOG_IMPORT")
        or os.environ.get("PREVIEW_CACHE_ONLY")
        or ""
    )
    return raw.strip().lower() in {"1", "true", "yes", "on"}

def _canonicalize_submap_name(parent_map: str, raw_submap: str) -> str:
    """Normalize imported submap text to a canonical submap for the parent map."""
    clean = (raw_submap or "").strip()
    if not clean:
        return ""

    candidates = MAP_SUBMAPS.get(parent_map, [])
    if not candidates:
        return clean

    clean_compact = _compact_text(clean)
    parent_compact = _compact_text(parent_map)

    # Some sheets repeat the parent map in the submap value.
    if parent_compact and clean_compact.startswith(parent_compact):
        clean_compact = clean_compact[len(parent_compact):]

    # Direct/fuzzy match against known submaps for this map.
    for candidate in candidates:
        candidate_compact = _compact_text(candidate)
        if clean_compact == candidate_compact:
            return candidate
        if clean_compact and (clean_compact in candidate_compact or candidate_compact in clean_compact):
            return candidate

    # Known analyst-sheet abbreviations.
    alias_lookup = {
        "Hell's Haven": {
            "fa": "Frozen Airfield",
            "ssf": "Super-Soldier Factory",
            "em": "Eldritch Monument",
            "supersoilderfactory": "Super-Soldier Factory",
            "hellshavensupersoilderfactory": "Super-Soldier Factory",
        },
        "Birin T'Challa": {
            "iis": "Imperial Institute of Science",
            "impinsscience": "Imperial Institute of Science",
            "ss": "Stellar Spaceport",
            "wf": "Warrior Falls",
        },
        "Krakoa": {
            "ca": "Carousel",
            "gr": "Grove",
            "cr": "Cradle",
        },
        "Celestial": {
            "co": "Codex",
            "va": "Vault",
            "ha": "Hand",
        },
    }

    mapped = alias_lookup.get(parent_map, {}).get(clean_compact)
    if mapped:
        return mapped

    # Fuzzy match for misspellings (e.g. Soilder vs Soldier).
    best_candidate = ""
    best_score = 0.0
    for candidate in candidates:
        candidate_compact = _compact_text(candidate)
        score = SequenceMatcher(None, clean_compact, candidate_compact).ratio()
        if score > best_score:
            best_score = score
            best_candidate = candidate
    if best_candidate and best_score >= 0.72:
        return best_candidate

    return clean


def split_score_pair(raw_score: str) -> tuple[str, str]:
    value = (raw_score or "").strip()
    if not value:
        return "", ""

    matches = re.findall(r"\d+(?:\.\d+)?", value)
    if not matches:
        return "", ""
    if len(matches) == 1:
        return matches[0], ""
    return matches[0], matches[1]


def build_score_text(left_score: str, right_score: str, fallback: str = "") -> str:
    left = (left_score or "").strip()
    right = (right_score or "").strip()
    if left and right:
        return f"{left}-{right}"
    if left:
        return left
    if right:
        return right
    return (fallback or "").strip()


def score_for_perspective(raw_score: str, *, perspective: str = "left") -> float | None:
    left_score, right_score = split_score_pair(raw_score)
    target = left_score if perspective == "left" else right_score
    if not target:
        return None
    try:
        return float(target)
    except ValueError:
        return None


def flip_score_text(raw_score: str) -> str:
    left_score, right_score = split_score_pair(raw_score)
    if left_score and right_score:
        return f"{right_score}-{left_score}"
    return (raw_score or "").strip()


def _our_team_slot_from_protect_order(raw_value: str) -> str:
    """
    Determine our team slot from the CSV 1st/2nd protect column.
    User rule: 2nd protect = team1, so 1st protect = team2.
    """
    value = (raw_value or "").strip().lower()
    if not value:
        return "team1"
    if value.startswith("2") or "second" in value:
        return "team1"
    if value.startswith("1") or "first" in value:
        return "team2"
    return "team1"


def _parse_csv_into_scrims(
    raw_text: str,
    team_id: int | None,
    team_name: str,
) -> tuple[list[dict], list[str]]:
    """
    Parse the analyst CSV format and return (scrim_list, warning_list).

    CSV structure (0-indexed columns):
      0   Date
      1   Enemy team name
      2   Map name (may include submap/side suffix or bracket hint)
      3   Our score
      4   Their score
      5   Result (Won/Lost)
      6   Ban Us1          7  Ban Them1
      8   Ban Us2          9  Ban Them2
      10  Ban Us3          11 Ban Them3
      12  Ban Us4          13 Ban Them4
      14  Protect Us1      15 Protect Them1
      16  Protect Us2      17 Protect Them2
      18  1st/2nd Protect  (ignored)
      19  (separator)
      20  Date (right half - used when left is blank)
      21  Enemy (right half)
      22  Map (right half)
      23  Us Result
      24-29  Our heroes (Tank,Tank,DPS,DPS,Supp,Supp)
      30  (separator)
      31  Them Result
      32-37  Their heroes
    """

    warnings: list[str] = []

    reader = csv.reader(io.StringIO(raw_text))
    rows = list(reader)
    if not rows:
        return [], ["CSV file is empty."]

    # Skip the header row (it starts with empty cells then "Ban Us1" etc.)
    header = rows[0]
    data_rows = rows[1:]

    def _pad(row: list[str]) -> list[str]:
        while len(row) < _CSV_MIN_COLS:
            row.append("")
        return row

    def _cell(row: list[str], idx: int) -> str:
        return row[idx].strip() if idx < len(row) else ""

    def _heroes(row: list[str], sl: slice) -> list[dict]:
        return [{"hero": normalize_hero_slot_value(h), "player": ""} for h in row[sl]]

    # Group rows into (date, enemy) buckets, keeping insertion order.
    # We preserve the order that maps appear so that sub-rows stay near parents.
    # Structure: { (date, enemy): [padded_row, ...] }
    from collections import OrderedDict
    buckets: OrderedDict[tuple[str, str], list[list[str]]] = OrderedDict()
    short_row_count = 0

    for raw_row in data_rows:
        if len(raw_row) < _CSV_MIN_COLS:
            short_row_count += 1
            continue

        row = _pad(list(raw_row))
        left_filled  = any(row[i].strip() for i in range(20))
        right_filled = any(row[i].strip() for i in range(20, _CSV_MIN_COLS))
        if not left_filled and not right_filled:
            continue

        if left_filled:
            date_val  = _cell(row, _CSV_DATE)
            enemy_val = _cell(row, _CSV_ENEMY)
        else:
            date_val  = _cell(row, _CSV_R_DATE)
            enemy_val = _cell(row, _CSV_R_ENEMY)

        if not date_val and not enemy_val:
            continue

        key = (date_val, enemy_val)
        buckets.setdefault(key, []).append(row)

    if not buckets:
        if short_row_count:
            return [], [f"Skipped {short_row_count} row(s) with missing columns."]
        return [], ["No data rows found in CSV."]

    # Build scrim objects from each bucket.
    all_scrims: list[dict] = []

    for (scrim_date, enemy_name), bucket_rows in buckets.items():
        maps: list[dict] = []
        current_parent_idx: int | None = None

        for row in bucket_rows:
            left_filled = any(row[i].strip() for i in range(20))

            # Gather map name from whichever side is available
            if left_filled:
                raw_map_name = _cell(row, _CSV_MAP)
            else:
                raw_map_name = _cell(row, _CSV_R_MAP)

            if not raw_map_name:
                continue

            normalized_map_name = raw_map_name.strip()
            normalized_map_name_lower = normalized_map_name.lower()
            is_side_row = normalized_map_name_lower.endswith(" attack") or normalized_map_name_lower.endswith(" defense")

            # Right-side heroes / results
            our_heroes  = _heroes(row, _CSV_R_US_H)
            their_heroes = _heroes(row, _CSV_R_TH_H)
            has_comp = any(h["hero"] for h in our_heroes + their_heroes)

            has_bans = any(
                row[i].strip()
                for i in [_CSV_BAN_US1, _CSV_BAN_TH1, _CSV_BAN_US2, _CSV_BAN_TH2,
                           _CSV_BAN_US3, _CSV_BAN_TH3, _CSV_BAN_US4, _CSV_BAN_TH4,
                           _CSV_PROT_US1, _CSV_PROT_TH1, _CSV_PROT_US2, _CSV_PROT_TH2]
            )

            # Parent map rows are map-level rows. Side split rows like "Midtown Attack"
            # should never become standalone parent maps.
            is_parent = has_bans or (left_filled and not has_comp and not is_side_row)

            if is_parent:
                canonical = _match_map_name(raw_map_name)
                base_name = _strip_bracket_hint(raw_map_name)
                our_team_slot = _our_team_slot_from_protect_order(_cell(row, _CSV_PROTECT_ORDER))
                enemy_team_slot = opposite_team_slot(our_team_slot)

                left_result = _cell(row, _CSV_RESULT)
                result_str  = "Win" if left_result == "Won" else "Loss" if left_result == "Lost" else ""

                score_us  = _cell(row, _CSV_SCORE_US)
                score_thm = _cell(row, _CSV_SCORE_THM)
                score_str = f"{score_us}-{score_thm}" if score_us or score_thm else ""

                draft_our = {
                    "ban1":     _cell(row, _CSV_BAN_US1),
                    "ban2":     _cell(row, _CSV_BAN_US2),
                    "protect1": _cell(row, _CSV_PROT_US1),
                    "ban3":     _cell(row, _CSV_BAN_US3),
                    "protect2": _cell(row, _CSV_PROT_US2),
                    "ban4":     _cell(row, _CSV_BAN_US4),
                }
                draft_enemy = {
                    "ban1":     _cell(row, _CSV_BAN_TH1),
                    "ban2":     _cell(row, _CSV_BAN_TH2),
                    "protect1": _cell(row, _CSV_PROT_TH1),
                    "ban3":     _cell(row, _CSV_BAN_TH3),
                    "protect2": _cell(row, _CSV_PROT_TH2),
                    "ban4":     _cell(row, _CSV_BAN_TH4),
                }
                draft = {
                    our_team_slot: draft_our,
                    enemy_team_slot: draft_enemy,
                }

                map_entry = {
                    "_base_name": base_name,        # temp field for sub-row matching
                    "map_name":   canonical,
                    "side":       "",
                    "our_team_slot": our_team_slot,
                    "result":     result_str,
                    "score":      score_str,
                    "draft":      draft,
                    "comp":       build_default_comp_sections(canonical),
                    "notes":      "",
                    "vod_url":    "",
                    "events":     [],
                }
                maps.append(map_entry)
                current_parent_idx = len(maps) - 1

            elif has_comp or is_side_row:
                # Sub-row: comp section belonging to the preceding parent
                # Determine parent by name prefix match
                parent_idx = current_parent_idx
                if parent_idx is None:
                    # No parent yet. Only create an implicit parent if this row has comp
                    # payload. For empty side rows, skip to avoid inflating map counts.
                    if not has_comp:
                        warnings.append(
                            f"Skipped side row without parent: {raw_map_name} ({scrim_date} vs {enemy_name})."
                        )
                        continue

                    # No parent yet - create an implicit parent
                    canonical = _match_map_name(raw_map_name)
                    map_entry = {
                        "_base_name": canonical,
                        "map_name":   canonical,
                        "side":       "",
                        "our_team_slot": "team1",
                        "result":     "",
                        "score":      "",
                        "draft":      {"team1": {"ban1":"","ban2":"","protect1":"","ban3":"","protect2":"","ban4":""}, "team2": {"ban1":"","ban2":"","protect1":"","ban3":"","protect2":"","ban4":""}},
                        "comp":       build_default_comp_sections(canonical),
                        "notes":      "",
                        "vod_url":    "",
                        "events":     [],
                    }
                    maps.append(map_entry)
                    parent_idx = len(maps) - 1
                    current_parent_idx = parent_idx

                parent = maps[parent_idx]
                base = parent["_base_name"]
                suffix = raw_map_name[len(base):].strip() if raw_map_name.lower().startswith(base.lower()) else raw_map_name

                submap = ""
                section_side = ""
                if suffix in ("Attack", "Defense"):
                    section_side = suffix
                elif suffix:
                    submap = _canonicalize_submap_name(parent.get("map_name", ""), suffix)

                score_us = _cell(row, _CSV_SCORE_US)
                score_thm = _cell(row, _CSV_SCORE_THM)
                section_score = ""
                if score_us and score_thm:
                    section_score = f"{score_us}-{score_thm}"
                elif score_us:
                    section_score = score_us
                elif score_thm:
                    section_score = score_thm

                # Fallback: some analyst sheets only include side outcome (Won/Lost)
                # for Attack/Defense rows. Use 1/0 so averages still populate.
                if not section_score:
                    us_result = _cell(row, _CSV_R_US_RES).lower()
                    if us_result == "won":
                        section_score = "1"
                    elif us_result == "lost":
                        section_score = "0"

                # Pad heroes to 6 slots each
                while len(our_heroes) < 6:
                    our_heroes.append({"hero": "", "player": ""})
                while len(their_heroes) < 6:
                    their_heroes.append({"hero": "", "player": ""})

                parent_our_slot = parent.get("our_team_slot", "team1")
                if parent_our_slot not in TEAM_SLOTS:
                    parent_our_slot = "team1"

                section = {
                    "submap": submap,
                    "side":   section_side,
                    "score":  section_score,
                    "team1":  our_heroes[:6] if parent_our_slot == "team1" else their_heroes[:6],
                    "team2":  our_heroes[:6] if parent_our_slot == "team2" else their_heroes[:6],
                }

                # Assign comps to the matching prebuilt section for this map type
                # (submap for Control, side for Escort/Hybrid). If no match exists,
                # append as a fallback.
                target_index = None
                for idx, existing in enumerate(parent.get("comp", [])):
                    existing_submap = (existing.get("submap") or "").strip().lower()
                    existing_side = (existing.get("side") or "").strip().lower()
                    if section_side:
                        if existing_side == section_side.strip().lower():
                            target_index = idx
                            break
                    elif submap:
                        if _compact_text(existing_submap) == _compact_text(submap):
                            target_index = idx
                            break

                if target_index is None:
                    parent.setdefault("comp", []).append(section)
                else:
                    parent["comp"][target_index].update(
                        {
                            "score": section_score,
                            "team1": our_heroes[:6] if parent_our_slot == "team1" else their_heroes[:6],
                            "team2": our_heroes[:6] if parent_our_slot == "team2" else their_heroes[:6],
                        }
                    )

        # Remove temp field
        for m in maps:
            m.pop("_base_name", None)

        if not maps:
            warnings.append(f"No maps found for scrim {scrim_date} vs {enemy_name}.")
            continue

        scrim = {
            "opponent":      enemy_name,
            "enemy_team":    enemy_name,
            "enemy_team_id": None,
            "scrim_date":    scrim_date,
            "season":       "",
            "team_id":       team_id,
            "team_name":     team_name,
            "notes":         "",
            "maps":          maps,
        }
        all_scrims.append(scrim)

    return all_scrims, warnings


def _parse_template_csv_into_scrims(
    raw_text: str,
    team_id: int | None,
    team_name: str,
) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []

    reader = csv.reader(io.StringIO(raw_text))
    rows = list(reader)
    if not rows:
        return [], ["CSV file is empty."]

    data_rows = rows[1:]
    selected_team_name = (team_name or "").strip()

    from collections import OrderedDict
    buckets: OrderedDict[tuple[str, str], list[list[str]]] = OrderedDict()

    def _pad(row: list[str]) -> list[str]:
        while len(row) < _TEMPLATE_CSV_MIN_COLS:
            row.append("")
        return row

    def _cell(row: list[str], idx: int) -> str:
        return row[idx].strip() if idx < len(row) else ""

    def _build_team_comp(row: list[str], pairs: tuple[tuple[int, int], ...]) -> list[dict]:
        slots: list[dict] = []
        for player_idx, hero_idx in pairs:
            raw_player_name = _cell(row, player_idx)
            slots.append(
                {
                    "player": "" if is_ringer_player_name(raw_player_name) else normalize_player_name(raw_player_name),
                    "hero": normalize_hero_slot_value(_cell(row, hero_idx)),
                }
            )
        return slots

    def _result_for_selected_team(result_value: str, team1_name: str, team2_name: str, our_slot: str) -> str:
        normalized = (result_value or "").strip().lower()
        if not normalized:
            return ""
        if normalized in {"draw", "tie"}:
            return ""
        if normalized in {"win", "won"}:
            return "Win"
        if normalized in {"loss", "lost"}:
            return "Loss"

        team1_normalized = (team1_name or "").strip().lower()
        team2_normalized = (team2_name or "").strip().lower()
        if normalized == team1_normalized:
            return "Win" if our_slot == "team1" else "Loss"
        if normalized == team2_normalized:
            return "Win" if our_slot == "team2" else "Loss"
        return ""

    for raw_row in data_rows:
        if not any(cell.strip() for cell in raw_row):
            continue
        row = _pad(list(raw_row))

        team1_name = _cell(row, _TEMPLATE_CSV_TEAM1)
        team2_name = _cell(row, _TEMPLATE_CSV_TEAM2)
        map_name = _cell(row, _TEMPLATE_CSV_MAP)
        scrim_date = _cell(row, _TEMPLATE_CSV_DATE)

        if not team1_name or not team2_name or not map_name or not scrim_date:
            continue

        if _team_names_match(team1_name, selected_team_name):
            enemy_name = team2_name
        elif _team_names_match(team2_name, selected_team_name):
            enemy_name = team1_name
        else:
            warnings.append(f"Skipped row for {team1_name} vs {team2_name} on {scrim_date}: selected team not found in row.")
            continue

        buckets.setdefault((scrim_date, enemy_name), []).append(row)

    if not buckets:
        return [], warnings or ["No matching rows found for the selected team in CSV."]

    all_scrims: list[dict] = []

    for (scrim_date, enemy_name), bucket_rows in buckets.items():
        maps: list[dict] = []
        for row in bucket_rows:
            team1_name = _cell(row, _TEMPLATE_CSV_TEAM1)
            team2_name = _cell(row, _TEMPLATE_CSV_TEAM2)
            canonical_map = _match_map_name(_cell(row, _TEMPLATE_CSV_MAP))
            our_team_slot = "team1" if _team_names_match(team1_name, selected_team_name) else "team2"

            team1_score = _cell(row, _TEMPLATE_CSV_SCORE_TEAM1)
            team2_score = _cell(row, _TEMPLATE_CSV_SCORE_TEAM2)
            score_text = build_score_text(team1_score, team2_score)
            result_text = infer_result_from_score_text(score_text, slot=our_team_slot)
            if not result_text:
                result_text = _result_for_selected_team(_cell(row, _TEMPLATE_CSV_RESULT), team1_name, team2_name, our_team_slot)

            draft = {
                "team1": {
                    "ban1": _cell(row, _TEMPLATE_CSV_TEAM1_BAN1),
                    "ban2": _cell(row, _TEMPLATE_CSV_TEAM1_BAN2),
                    "protect1": _cell(row, _TEMPLATE_CSV_TEAM1_SAVE1),
                    "ban3": _cell(row, _TEMPLATE_CSV_TEAM1_BAN3),
                    "protect2": _cell(row, _TEMPLATE_CSV_TEAM1_SAVE2),
                    "ban4": _cell(row, _TEMPLATE_CSV_TEAM1_BAN4),
                },
                "team2": {
                    "ban1": _cell(row, _TEMPLATE_CSV_TEAM2_BAN1),
                    "ban2": _cell(row, _TEMPLATE_CSV_TEAM2_BAN2),
                    "protect1": _cell(row, _TEMPLATE_CSV_TEAM2_SAVE1),
                    "ban3": _cell(row, _TEMPLATE_CSV_TEAM2_BAN3),
                    "protect2": _cell(row, _TEMPLATE_CSV_TEAM2_SAVE2),
                    "ban4": _cell(row, _TEMPLATE_CSV_TEAM2_BAN4),
                },
            }

            comp_sections = build_default_comp_sections(canonical_map)
            if comp_sections:
                team1_comp = _build_team_comp(row, _TEMPLATE_CSV_TEAM1_PLAYERS)
                team2_comp = _build_team_comp(row, _TEMPLATE_CSV_TEAM2_PLAYERS)
                for idx, section in enumerate(comp_sections):
                    section["team1"] = copy.deepcopy(team1_comp)
                    section["team2"] = copy.deepcopy(team2_comp)
                    if idx == 0:
                        section["score"] = score_text

            maps.append(
                {
                    "map_name": canonical_map,
                    "map_type": normalize_map_type_value(_cell(row, _TEMPLATE_CSV_MAP_TYPE)),
                    "side": "",
                    "our_team_slot": our_team_slot,
                    "result": result_text,
                    "score": score_text,
                    "draft": draft,
                    "comp": comp_sections,
                    "notes": _cell(row, _TEMPLATE_CSV_NOTE),
                    "vod_url": "",
                    "events": [],
                    "team1_name": team1_name,
                    "team2_name": team2_name,
                }
            )

        if not maps:
            continue

        all_scrims.append(
            {
                "opponent": enemy_name,
                "enemy_team": enemy_name,
                "enemy_team_id": None,
                "scrim_date": scrim_date,
                "season": "",
                "team_id": team_id,
                "team_name": team_name,
                "notes": "",
                "maps": maps,
            }
        )

    return all_scrims, warnings


def parse_csv_into_scrims(
    raw_text: str,
    team_id: int | None,
    team_name: str,
) -> tuple[list[dict], list[str]]:
    reader = csv.reader(io.StringIO(raw_text))
    rows = list(reader)
    if not rows:
        return [], ["CSV file is empty."]

    header = [cell.strip().lower() for cell in rows[0]]
    is_template_csv = (
        len(header) >= _TEMPLATE_CSV_MIN_COLS
        and "replay code" in header
        and "map type" in header
        and "team 1" in header
        and "team 2" in header
        and "map" in header
    )

    if is_template_csv:
        return _parse_template_csv_into_scrims(raw_text, team_id, team_name)
    return _parse_csv_into_scrims(raw_text, team_id, team_name)


def summarize_import_warnings(warnings: list[str], *, preview_count: int = 5) -> str:
    if not warnings:
        return ""
    preview = warnings[:preview_count]
    remaining = len(warnings) - len(preview)
    message = " | ".join(preview)
    if remaining > 0:
        message += f" | {remaining} more warning(s)"
    return message


@app.route("/scrims/import-csv", methods=["POST"])
def import_csv_scrims():
    global NEXT_SCRIM_ID

    team_id  = parse_team_id(request.form.get("team_id", ""))
    team_name = get_team_name_by_id(team_id)
    season = normalize_season_value(request.form.get("season", ""))
    if not team_name:
        flash("Please select your team before importing.", "error")
        return redirect(url_for("scrims"))
    if not season:
        flash("Please set a season for this import.", "error")
        return redirect(url_for("scrims"))

    file = request.files.get("csv_file")
    if not file or not file.filename:
        flash("No CSV file selected.", "error")
        return redirect(url_for("scrims"))

    ext = Path(file.filename).suffix.lower()
    if ext not in {".csv", ".txt", ".xlsx"}:
        flash("Only .csv, .txt, or .xlsx files are supported.", "error")
        return redirect(url_for("scrims"))

    try:
        file_bytes = file.read()
        if ext == ".xlsx":
            raw_text = _extract_csv_text_from_xlsx(file_bytes)
        else:
            raw_text = _parse_text_payload(file_bytes)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("scrims"))
    except Exception:
        flash("Could not parse the import file. Make sure the file is a valid CSV/TXT/XLSX export.", "error")
        return redirect(url_for("scrims"))

    parsed_scrims, warnings = parse_csv_into_scrims(raw_text, team_id, team_name)

    if not parsed_scrims:
        warning_summary = summarize_import_warnings(warnings)
        flash("No scrims could be imported from that CSV. " + warning_summary, "error")
        return redirect(url_for("scrims"))

    # Try to match enemy team names to existing teams in the global team database
    db = get_db()
    migrate_enemy_teams_to_team_database(db)
    enemy_rows = db.execute(
        "SELECT id, name FROM teams WHERE id != ?", (team_id,)
    ).fetchall() if team_id else []
    enemy_lookup: dict[str, int] = {}
    for row in enemy_rows:
        for key in _team_name_match_keys(row["name"]):
            enemy_lookup.setdefault(key, row["id"])

    imported = 0
    updated = 0
    for scrim in parsed_scrims:
        scrim["season"] = season
        normalize_scrim_record(scrim)
        _prepare_imported_scrim_context(scrim, team_id, team_name, enemy_lookup)
        _sync_scrim_rosters_with_database(scrim)

        if scrim.get("enemy_team_id"):
            for enemy_key in _team_name_match_keys(scrim["enemy_team"]):
                enemy_lookup[enemy_key] = scrim["enemy_team_id"]

        existing_scrim = _find_duplicate_scrim_for_import(scrim)
        if existing_scrim is not None:
            _merge_imported_scrim(existing_scrim, scrim)
            _assign_missing_scrim_ids(existing_scrim)
            updated += 1
            continue

        scrim["id"] = NEXT_SCRIM_ID
        NEXT_SCRIM_ID += 1
        _assign_missing_scrim_ids(scrim)
        SCRIMS.append(scrim)
        imported += 1

    save_app_state()

    msg_parts = []
    if imported:
        msg_parts.append(f"Imported {imported} scrim{'s' if imported != 1 else ''}")
    if updated:
        msg_parts.append(f"updated {updated} duplicate{'s' if updated != 1 else ''}")
    msg = ". ".join(msg_parts) + "."
    if warnings:
        msg += " Warnings: " + summarize_import_warnings(warnings)
    flash(msg, "success")
    return redirect(url_for("scrims"))


@app.route("/scrims/import-log", methods=["POST"])
def import_scrimcore_log_scrim():
    global NEXT_SCRIM_ID

    team_id = parse_team_id(request.form.get("team_id", ""))
    team_name = get_team_name_by_id(team_id)
    season = normalize_season_value(request.form.get("season", ""))
    if not team_name:
        flash("Please select your team before importing a log.", "error")
        return redirect(url_for("scrims"))
    if not season:
        flash("Please set a season for this import.", "error")
        return redirect(url_for("scrims"))

    scrim_date = (request.form.get("scrim_date", "") or "").strip() or date.today().isoformat()
    enemy_name_input = (request.form.get("enemy_name", "") or "").strip()

    files = [file for file in request.files.getlist("log_file") if file and file.filename]
    if not files:
        flash("No log file selected.", "error")
        return redirect(url_for("scrims"))

    db = get_db()
    migrate_enemy_teams_to_team_database(db)
    enemy_rows = db.execute(
        "SELECT id, name FROM teams WHERE id != ?", (team_id,)
    ).fetchall() if team_id else []
    enemy_lookup: dict[str, int] = {}
    for row in enemy_rows:
        for key in _team_name_match_keys(row["name"]):
            enemy_lookup.setdefault(key, row["id"])

    cache_only_mode = _cache_only_log_import_mode()
    imported = 0
    updated = 0
    skipped: list[str] = []
    last_scrim_id = None

    for file in files:
        ext = Path(file.filename).suffix.lower()
        if ext not in {".txt", ".csv", ".log", ".xlsx"}:
            skipped.append(f"{file.filename}: unsupported file type")
            continue

        try:
            file_bytes = file.read()
            raw_text = _extract_csv_text_from_xlsx(file_bytes) if ext == ".xlsx" else _parse_text_payload(file_bytes)
        except ValueError as exc:
            skipped.append(f"{file.filename}: {exc}")
            continue
        except Exception:
            skipped.append(f"{file.filename}: could not read file")
            continue

        parsed_rows, parser_summary = parse_scrimcore_log_text(raw_text)
        if not parsed_rows:
            skipped.append(f"{file.filename}: no parseable events")
            continue

        parsed_team1 = (parser_summary.get("team1_name") or "").strip()
        parsed_team2 = (parser_summary.get("team2_name") or "").strip()
        our_team_slot = "team1"
        if parsed_team2 and _team_names_match(parsed_team2, team_name):
            our_team_slot = "team2"

        inferred_enemy = enemy_name_input
        if not inferred_enemy:
            inferred_enemy = (parsed_team2 if our_team_slot == "team1" else parsed_team1) or "Unknown Opponent"
        if _is_scrimcore_generic_team_name(inferred_enemy):
            inferred_enemy = "Opponent"

        parsed_map_name = (parser_summary.get("map_name") or "").strip()
        map_name = _match_map_name(parsed_map_name) if parsed_map_name else "Unknown Map"
        players_by_label = _extract_scrimcore_players_by_team(parsed_rows)
        draft_data = _extract_scrimcore_draft(parsed_rows, parsed_team1, parsed_team2, our_team_slot)
        comp_sections = _extract_scrimcore_comp_sections(parsed_rows, parsed_team1, parsed_team2, our_team_slot, map_name)
        timeline_events, killfeed_summary = _build_scrimcore_killfeed_events(parsed_rows, parsed_team1, parsed_team2, our_team_slot, map_name)
        analysis_events, analysis_summary = _build_scrimcore_analysis_events(parsed_rows, parsed_team1, parsed_team2, our_team_slot, map_name)
        map_score, map_result, _round_scores = _extract_scrimcore_round_scores(parsed_rows, our_team_slot)
        parser_summary["killfeed"] = killfeed_summary
        parser_summary["analysis"] = analysis_summary
        parser_summary["source_file"] = file.filename
        if map_score:
            parser_summary["score"] = map_score
        if map_result:
            parser_summary["result"] = map_result

        map_entry = {
            "map_name": map_name,
            "map_type": DEFAULT_MAP_TYPE,
            "side": "",
            "our_team_slot": our_team_slot,
            "result": map_result,
            "score": map_score,
            "draft": draft_data,
            "comp": comp_sections or build_default_comp_sections(map_name),
            "notes": "",
            "vod_url": "",
            "events": sorted(
                [*timeline_events, *analysis_events],
                key=lambda item: (
                    float(item.get("match_time") or 0),
                    1 if item.get("event_source") == "scrimcore_first_death_marker" else 0,
                    str(item.get("event_source") or ""),
                ),
            ),
            "round_hero_summary": _build_round_hero_summary(comp_sections),
            "parser_summary": parser_summary,
            "parser_source": "scrimcore-log-killfeed-import",
            "team1_name": team_name if our_team_slot == "team1" else inferred_enemy,
            "team2_name": inferred_enemy if our_team_slot == "team1" else team_name,
        }

        scrim = {
            "opponent": inferred_enemy,
            "enemy_team": inferred_enemy,
            "enemy_team_id": None,
            "scrim_date": scrim_date,
            "season": season,
            "team_id": team_id,
            "team_name": team_name,
            "team_slot": our_team_slot,
            "team1_players": _resolve_scrimcore_players_for_team(
                players_by_label,
                team_name if our_team_slot == "team1" else inferred_enemy,
                [parsed_team1 if our_team_slot == "team1" else parsed_team2],
            ),
            "team2_players": _resolve_scrimcore_players_for_team(
                players_by_label,
                inferred_enemy if our_team_slot == "team1" else team_name,
                [parsed_team2 if our_team_slot == "team1" else parsed_team1],
            ),
            "notes": f"Imported from ScrimCore-style raw log parser. Parsed {parser_summary.get('total_rows', 0)} rows across {parser_summary.get('event_type_count', 0)} event types.",
            "maps": [map_entry],
        }

        normalize_scrim_record(scrim)
        _prepare_imported_scrim_context(scrim, team_id, team_name, enemy_lookup)
        _sync_scrim_rosters_with_database(scrim)

        existing_scrim = _find_duplicate_scrim_for_import(scrim)
        if existing_scrim is not None:
            _merge_imported_scrim(existing_scrim, scrim)
            _assign_missing_scrim_ids(existing_scrim)
            updated += 1
            last_scrim_id = existing_scrim.get("id")
            continue

        scrim["id"] = NEXT_SCRIM_ID
        NEXT_SCRIM_ID += 1
        _assign_missing_scrim_ids(scrim)
        SCRIMS.append(scrim)
        imported += 1
        last_scrim_id = scrim["id"]

    if not cache_only_mode:
        save_app_state()

    if not imported and not updated:
        flash("No logs could be imported. " + summarize_import_warnings(skipped), "error")
        return redirect(url_for("scrims"))

    msg_parts = []
    if imported:
        msg_parts.append(f"Imported {imported} log{'s' if imported != 1 else ''}")
    if updated:
        msg_parts.append(f"updated {updated} duplicate{'s' if updated != 1 else ''}")
    msg = ". ".join(msg_parts) + "."
    if skipped:
        msg += " Warnings: " + summarize_import_warnings(skipped)
    if cache_only_mode:
        msg += " Preview cache-only mode: not saved to DB."
    flash(msg, "success")

    if len(files) == 1 and last_scrim_id:
        return redirect(url_for("scrim_detail", scrim_id=last_scrim_id))
    return redirect(url_for("scrims"))


